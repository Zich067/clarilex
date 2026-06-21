"""產生「法律專家標註表」Excel:給獨立、有法律背景的人填 supported/partial/unsupported。

刻意隱藏 mode(baseline/rag/triangulation)與評審原判定,避免偏誤。
完成回收後,以 sample_id 對回 judge_validation_key.json 計算 judge–human κ。

用法:PYTHONPATH=. .venv/bin/python scripts/build_expert_annotation_xlsx.py
輸出:data/results/expert_annotation_sheet.xlsx
"""
import csv
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"

RELEVANT_LAWS = ["民法", "消費者保護法", "民事訴訟法"]


def norm_no(s: str) -> str:
    return re.sub(r"[^\d\-]", "", s)


def build_article_map():
    raw = json.load(open(DATA / "laws" / "ChLaw.json", encoding="utf-8-sig"))
    by_no = {}
    for law in raw["Laws"]:
        name = law["LawName"]
        if name not in RELEVANT_LAWS:
            continue
        for a in law.get("LawArticles", []):
            no = norm_no(a.get("ArticleNo", ""))
            content = (a.get("ArticleContent") or "").strip()
            if no:
                by_no.setdefault(no, []).append((name, content))
    return by_no


def squash(s: str) -> str:
    return re.sub(r"\s+", "", s or "")


def resolve_refs(ref_field: str, by_no):
    """把 reference_articles 欄拆成各條,並抓回完整法條原文。"""
    out = []
    # 以「第 X 條:」為界切割
    parts = re.split(r"；|;", ref_field)
    for part in parts:
        m = re.match(r"\s*第\s*([\d\-]+)\s*條\s*[:：]?\s*(.*)", part, re.S)
        if not m:
            continue
        no = norm_no(m.group(1))
        trunc = squash(m.group(2))[:20]
        candidates = by_no.get(no, [])
        chosen = None
        if len(candidates) == 1:
            chosen = candidates[0]
        elif candidates:
            for name, content in candidates:
                if trunc and squash(content).startswith(trunc):
                    chosen = (name, content)
                    break
            if chosen is None:
                chosen = candidates[0]
        if chosen:
            name, content = chosen
            out.append(f"{name} 第 {no} 條：{content}")
        else:
            out.append(f"第 {no} 條:(原文待補)")
    return out


def main():
    gold = {}
    with open(DATA / "gold" / "lease_sale_gold.jsonl", encoding="utf-8") as f:
        for line in f:
            if line.strip():
                r = json.loads(line)
                gold[r["id"]] = r

    rows = list(csv.DictReader(open(DATA / "results" / "judge_validation_sheet.csv", encoding="utf-8-sig")))

    wb = Workbook()

    # ---------- 說明 sheet ----------
    ws0 = wb.active
    ws0.title = "填寫說明"
    instructions = [
        ("法律 RAG 系統輸出之主張稽核 — 專家標註表", True, 16),
        ("", False, 11),
        ("【目的】", True, 12),
        ("這是一份碩士論文(臺灣租賃/買賣民事領域之 AI 法律文件分析系統)的驗證資料。", False, 11),
        ("系統會把產生的法律分析報告拆成一條條「主張」,並各自引用法條。我們想請您這位具法律背景的", False, 11),
        ("獨立第三方,判斷每一條主張是否真的能由它所引用的法條獲得支持。", False, 11),
        ("", False, 11),
        ("【您要做的事】", True, 12),
        ("到「標註表」分頁,逐列閱讀「系統主張」與「引用之法條原文」,在『判定』欄的下拉選單中選一項。", False, 11),
        ("若願意,請在『理由與建議』欄簡述判斷依據(非必填,但對研究幫助很大)。共 60 列。", False, 11),
        ("", False, 11),
        ("【三個判定的定義】", True, 12),
        ("supported(支持):完整全對——主張的內容與要件,都能在引用法條中找到對應。", False, 11),
        ("partial(部分支持):對一半——方向正確,但有要件偏差、過度延伸,或只對應到部分條文。", False, 11),
        ("unsupported(不支持):整個不對——主張無法由引用法條支持(條文不對應、要件錯誤,或無中生有)。", False, 11),
        ("", False, 11),
        ("【重要說明】", True, 12),
        ("1. 每條主張由哪一種系統產生,已刻意隱藏,以避免影響您的判斷——請僅就「主張 vs 引用法條」評估。", False, 11),
        ("2. 判斷基準是「這條主張能否在它所引用的法條裡找到支持」,而非主張在法律上是否最完整正確。", False, 11),
        ("3. 法條原文取自全國法規資料庫(民法/消費者保護法)。", False, 11),
        ("4. 完成後請整份回傳即可。非常感謝您的協助!", False, 11),
    ]
    for i, (text, bold, size) in enumerate(instructions, start=1):
        c = ws0.cell(row=i, column=1, value=text)
        c.font = Font(bold=bold, size=size)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws0.column_dimensions["A"].width = 100

    # ---------- 標註表 sheet ----------
    ws = wb.create_sheet("標註表")
    headers = ["編號", "案件情境", "系統主張", "引用之法條原文", "判定 (supported / partial / unsupported)", "理由與建議 (非必填)"]
    header_fill = PatternFill("solid", fgColor="4472C4")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = header_fill
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = border

    for i, r in enumerate(rows, start=2):
        g = gold.get(r["gold_id"], {})
        query = g.get("query", "")
        ws.cell(row=i, column=1, value=r["sample_id"])
        ws.cell(row=i, column=2, value=query)
        ws.cell(row=i, column=3, value=r["claim"])
        ws.cell(row=i, column=4, value=r["reference_articles"])  # sheet 已含 shared 參照全文
        ws.cell(row=i, column=5, value="")
        ws.cell(row=i, column=6, value="")
        for col in range(1, 7):
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
        # 待填欄淡黃底提示
        for col in (5, 6):
            ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor="FFF2CC")

    widths = {1: 8, 2: 30, 3: 42, 4: 60, 5: 22, 6: 30}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 40

    dv = DataValidation(
        type="list",
        formula1='"supported,partial,unsupported"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "請從 supported / partial / unsupported 三者擇一"
    dv.errorTitle = "無效輸入"
    dv.prompt = "請下拉選擇"
    ws.add_data_validation(dv)
    dv.add(f"E2:E{len(rows)+1}")

    out = DATA / "results" / "expert_annotation_sheet.xlsx"
    wb.save(out)
    print(f"[OK] 已產生 {out}")
    print(f"     共 {len(rows)} 列待標註;判定欄已設下拉(supported/partial/unsupported)")


if __name__ == "__main__":
    main()
